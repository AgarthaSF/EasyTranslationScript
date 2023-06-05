# -*- coding: utf-8 -*-
import sys
import click
import uuid
import requests
import hashlib
import time
import json
# from imp import reload
import time
from openpyxl import Workbook
from openpyxl import load_workbook

# reload(sys)

YOUDAO_URL = 'https://openapi.youdao.com/api'
# replace them with your own app key and secret
APP_KEY = '${Your_APP_KEY}'
APP_SECRET = '${Your_APP_SECRET}'


def encrypt(signStr):
    hash_algorithm = hashlib.sha256()
    hash_algorithm.update(signStr.encode('utf-8'))
    return hash_algorithm.hexdigest()


def truncate(q):
    if q is None:
        return None
    size = len(q)
    return q if size <= 20 else q[0:10] + str(size) + q[size - 10:size]


def do_request(data):
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    return requests.post(YOUDAO_URL, data=data, headers=headers)


def connect(query):
    q = query
    data = {'from': 'en', 'to': 'zh-CHS', 'signType': 'v3'}
    curtime = str(int(time.time()))
    data['curtime'] = curtime
    salt = str(uuid.uuid1())
    signStr = APP_KEY + truncate(q) + salt + curtime + APP_SECRET
    sign = encrypt(signStr)
    data['appKey'] = APP_KEY
    data['q'] = q
    data['salt'] = salt
    data['sign'] = sign

    response = do_request(data)
    if response.status_code != 200:
        return
    else:
        content = response.content
        content_obj = json.loads(content)

        usPhonetic = ''
        explains = []
        if('basic' in content_obj):
            basic = content_obj['basic']
            if('us-phonetic' in basic):
                usPhonetic = basic['us-phonetic']
                usPhonetic = '[' + usPhonetic + ']'
            if('explains' in basic):
                explains = basic['explains']
        return usPhonetic, explains


def read_txt(file_name):
    wordList = []
    with open(file_name, "r") as f:
        for line in f.readlines():
            word = line.strip('\n').strip()
            if word != "":
                wordList.append(word)
    return wordList


def create_excel(fileName, wordList):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Word'
    ws['B1'] = 'Phonetic'
    ws['C1'] = 'Explanation'
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    for word in wordList:
        phonetic, explains = connect(word)
        expStr = ' '.join(explains)
        newRow = [word, phonetic, expStr]
        ws.append(newRow)
        time.sleep(0.8)
    wb.save(fileName)

@click.command()
@click.option('--input', '-i', default='', help='inputFileName')
@click.option('--output', '-o', default='', help='outputFileName')
def start(input, output):
    if input == '':
        print("argument missed, please add  -i <inputfile>")
        sys.exit()

    if output == '':
        output = input.split('.')[0] + '.xlsx'

    print(input, output)
    wordList = read_txt(input)
    create_excel(output, wordList)
    print('translation finished')

if __name__ == '__main__':
    start()